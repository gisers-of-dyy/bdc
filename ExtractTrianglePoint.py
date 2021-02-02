import re
import csv
from openpyxl import load_workbook
import xlwt
import random

import re
import csv
from openpyxl import load_workbook
import xlwt
import random
import ogr,os,osr,sys
from math import ceil

try:
    from osgeo import gdal
    from osgeo import ogr
except ImportError:
    import gdal
    import ogr
def get_Location(locationFileName):
    #一、打开文件
    f = open(locationFileName,encoding="UTF-8")
    locationContext = f.read()
    splitLocationContext = re.split(r'}}', locationContext)  # 使用 }}作为字符串的分隔符
    index = 0
    while index <len(splitLocationContext):
        #提取一行
        oneRow = splitLocationContext[index]
        # 每一行按照逗号分隔
        splitOneRow = re.split(r',', oneRow)
        tempId = 0
        tempList = []
        while tempId < len(splitOneRow):
            if splitOneRow[tempId].rfind(':') != -1:
                startIndex = splitOneRow[tempId].rfind(':')
                if splitOneRow[tempId].rfind('FRUITID') != -1:
                    value = splitOneRow[tempId][startIndex + 1:len(splitOneRow[tempId])-1]
                    tempList.append(value)
                else:
                    value = splitOneRow[tempId][startIndex + 1:]
                    tempList.append(value)
                tempId += 1
            else:
                tempId += 1

        if len(tempList)==3:
            PointPostionDict = dict(fruitId=(tempList[0]), x=tempList[1], y=tempList[2])
            #print(PointPostionDict)
            locationList.append(PointPostionDict)
        index+=1
        f.close()

def get_Attribute(attributeFileName):
    #一、打开文件
    f = open(attributeFileName,encoding="UTF-8")
    rowIndex = 0
    #按照行进行读取
    fruitIdTemp= -1
    fruitCategoryIdTemp = -1
    tempList = []
    while True:
        #5行为一循环
        line = f.readline()
        #获取fruitCategoryIdd的值
        if line.find('fruitCategoryId') != -1:
            fruitCategoryIndex = line.find(':')
            fruitCategoryIdTemp = line[fruitCategoryIndex+1 :len(line)-2 ]
            rowIndex +=1
        #获取 fruitIdIndex
        elif line.find('fruitId') != -1:
            fruitIdIndex = line.find(':')
            fruitIdTemp = line[fruitIdIndex+1 :len(line)-1]
            rowIndex += 1
        #获取 fruitAttrList
        elif line.find('fruitAttrList') != -1:
            # 每一行按照逗号分隔
            splitLine = re.split(r'}', line)  # 使用attrName 作为字符串的分隔符
            rowIndex += 1
            #获取点名
            temp = 0
            while temp < len(splitLine):
                tempIndex = splitLine[temp].find("attrValue")
                if tempIndex != -1:
                    startIndex= splitLine[temp].rfind(":")
                    value = splitLine[temp][startIndex + 3 :len(splitLine[temp])- 2 ]

                    tempList.append(value)
                    temp += 1
                else:
                    temp += 1
        #{ 或者}， 跳过
        else:
            rowIndex += 1

        if (rowIndex % 5 == 0):
            PointAttribute = dict(fruitId= fruitIdTemp, fruitCategoryId = fruitCategoryIdTemp,
                              pointName= tempList[0], newMapNumber= tempList[1], pointLeveld= tempList[2],
                              geoDatum = tempList[3],producDate = tempList[4])
            attributeList.append(PointAttribute)
            #print(PointAttribute)
            fruitIdTemp = -1
            fruitCategoryIdTemp = -1
            tempList.clear()

        if not line:
            break

def writer(outFileName):
    book = xlwt.Workbook()
    book_name = outFileName
    # 创建表单
    book.encoding = "utf-8"  # Equals to workbook = xlwt.Workbook(encoding="utf-8")
    newSheet = book.add_sheet(sheetname='三角点整理', cell_overwrite_ok=True)
    style = xlwt.XFStyle()  # 初始化样式
    font = xlwt.Font()  # 为样式创建字体
    font.name = '宋体'
    font.bold = True
    style.font = font  # 为样式设置字体

    column0 = ["fruitId", "x", "y", "fruitCategoryId", "点名",  "新图号", "等级","大地基准","生产时间"]
    k = 0
    for k in range(0, len(column0)):
        newSheet.write(0, k, column0[k], style)

    if len(locationList) == len(attributeList):
        row = 0
        while row < len(locationList):
            #按行处理
            dic = locationList[row]
            attribute = attributeList[row]
            for key in attribute:
                if dic.get(key):
                    pass
                else:
                    dic[key] = attribute[key]
            #每一行的每一个元素
            col = 0
            for jj in dic.values():
                newSheet.write(row + 1, col, jj, style)
                col+=1

            row += 1
    book.save(book_name)  # 保存工作簿
    print(f"{book_name}表格写入数据成功！")

def writeShp(outFileName):
    # 为了支持中文路径，请添加下面这句代码
    # gdal.SetConfigOption("GDAL_FILENAME_IS_UTF8", "NO")
    # # 为了使属性表字段支持中文，请添加下面这句
    # gdal.SetConfigOption("SHAPE_ENCODING", "")
    gdal.SetConfigOption("GDAL_FILENAME_IS_UTF8", "YES")
    gdal.SetConfigOption("SHAPE_ENCODING", "GBK")
    # 1.创建输出文件
    srs = osr.SpatialReference()  # 创建空间参考
    srs.ImportFromEPSG(4490)  # 定义地理坐标系CGCS2000
    outdriver = ogr.GetDriverByName('ESRI Shapefile')
    if os.path.exists(outFileName):
        outdriver.DeleteDataSource(outFileName)
    outds = outdriver.CreateDataSource(outFileName)
    oLayer = outds.CreateLayer(outFileName,srs,geom_type=ogr.wkbPoint)
    if oLayer == None:
        print("图层创建失败！\n")
        return

    # 2.下面创建属性表
    # 2.1 创建一个叫FieldID的整型属性
    fruitId = ogr.FieldDefn("fruitId", ogr.OFTInteger)
    oLayer.CreateField(fruitId, 1)
    #  2.2 创建一个叫坐标X的浮点属性
    x_coord = ogr.FieldDefn("x_coord", ogr.OFTString)
    oLayer.CreateField(x_coord, 1)
    #  2.3 创建一个叫坐标y的浮点属性
    y_coord = ogr.FieldDefn("y_coord", ogr.OFTString)
    oLayer.CreateField(y_coord, 1)
    #  2.4 创建一个叫fruitCategoryId的整型属性
    CategoryId = ogr.FieldDefn("CategoryId", ogr.OFTInteger)
    oLayer.CreateField(CategoryId, 1)

    #  2.5 创建一个叫pointName的字符型属性
    pointName = ogr.FieldDefn("pointName", ogr.OFTString)
    #oFieldName.SetWidth(100)
    oLayer.CreateField(pointName, 1)

    newMapNum = ogr.FieldDefn("newMapNum", ogr.OFTString)
    newMapNum.SetWidth(30)
    oLayer.CreateField(newMapNum, 1)

    #  2.6 创建一个叫pointLevel的字符型属性
    pointLevel = ogr.FieldDefn("pointLevel", ogr.OFTString)
    oLayer.CreateField(pointLevel, 1)

    #   2.7 创建一个叫geoDatum的字符型属性
    geoDatum = ogr.FieldDefn("geoDatum", ogr.OFTString)
    oLayer.CreateField(geoDatum, 1)

    #   2.8 创建一个叫producDate的字符型属性
    producDate = ogr.FieldDefn("producDate", ogr.OFTString)
    oLayer.CreateField(producDate, 1)

    #3.写入属性
    outfielddefn = oLayer.GetLayerDefn()
    if len(locationList) == len(attributeList):
        row = 0
        while row < len(locationList):
            # 按行处理
            dic = locationList[row]
            #str = attributeList[row].get('geoDatum')
            str = attributeList[row]['geoDatum']
            #print(row,str)
            if str == '2000国家大地坐标系':
                attribute = attributeList[row]
                for key in attribute:
                    if dic.get(key):
                        pass
                    else:
                        dic[key] = attribute[key]
                # 每一行的每一个元素
                #print(dic)
                # 创建点属性
                outfeat = ogr.Feature(outfielddefn)
                #添加字段属性
                col = 0
                for jj in dic.values():
                    outfeat.SetField(col, jj)
                    col+=1
                    # 创建要素，写入多边形
                point = ogr.Geometry(ogr.wkbPoint)
                # 构建几何类型:点
                point_x = float(dic['x'])
                point_y = float(dic['y'])
                point.AddPoint(point_x, point_y)  # 创建点

                outfeat.SetGeometry(point)
                # 写入图层
                oLayer.CreateFeature(outfeat)
                #oFeaturePint = None
            row += 1


locationFileName = "D:\Python\ExtractPointInfo\三角点\三角点位置.txt"
attributeFileName = "D:\Python\ExtractPointInfo\三角点\三角点属性.txt"
outFileName = "D:\Python\ExtractPointInfo\三角点\三角点整理.xls"

locationList = []
get_Location(locationFileName)

attributeList = []
get_Attribute(attributeFileName)

#两个字典排序按照fruitId排序
foo = lambda s:s['fruitId']
locationList.sort(key = foo)
attributeList.sort(key = foo)


#writer(outFileName)
outFileName1 = "D:\Python\ExtractPointInfo\三角点\TrianglePoint.shp"
writeShp(outFileName1)




# PointAttribute = dict(fruitId = -1,fruitCategoryId = -1, pointName ='xx', pointId='xx', pointLeveld='xx',
#               geoDatum = 'xx',heightDatum ='xx',producDate = 'xx')
# print(PointAttribute)    # {'name': '王大锤', 'age': 55, 'weight': 60, 'home': '中同仁路8号'}
#
# PointPostion = dict(fruitId = -1,x = -9999.0, y = -9999.0)


