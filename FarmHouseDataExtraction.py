# import os
#
# root = 'G:\潼南农房数据清理\原始数据\ftpfjyj'
# # path = os.path.join(root, 'part_A_final', 'train_data', 'images')
# pathnames = []
# for (dirpath, dirnames, filenames) in os.walk(root):
#     for dirnames in dirnames:
#         pathnames += [os.path.join(dirpath, dirnames)]
#
# print(pathnames)
#


import datetime
import os
from openpyxl import load_workbook
import xlwt
import os
import shutil
import re

class ExcelData(object):
    def __init__(self):
        #存储信息  --列表
        self.newHolder = []
        self.origiHolderList = []

    def segHolders(self,holdString):
        strLen = len(holdString)
        #1.查找有几个冒号
        colonNum = holdString.count(':')
        #print(holdString)

        # 2 切片分割
        if  colonNum == 1:
            #冒号个数为1 只有现权利人 切片冒号-字符串末尾
            conpileText = re.compile('现所有权人:')
            newHolder = conpileText.sub('', holdString)
            self.newHolder = newHolder
        else:
            # 冒号个数为多个  有现权利人和原权利人（个数未知）
            splitString = re.split(r',', holdString)  # 使用 , 作为字符串的分隔符
            if len(splitString)>1:
                conpileText1 = re.compile('现所有权人:')
                newHolder = conpileText1.sub('', splitString[0])
                self.newHolder = newHolder
                origiNum = 1
                origiHolderList = []
                while origiNum < len(splitString):
                    conpileText2 = re.compile('原所有权人:')
                    origiHolder = conpileText2.sub('', splitString[origiNum])
                    origiHolderList.append(origiHolder)
                    origiNum += 1
                self.origiHolderList = origiHolderList

def copy_dirs(src_path, target_path):
    # copy     all    # files    # of    # src_path    # to    # target_path’’’

    file_count = 0
    source_path = os.path.abspath(src_path)
    target_path = os.path.abspath(target_path)
    if not os.path.exists(target_path):
        os.makedirs(target_path)
    if os.path.exists(source_path):
        for root, dirs, files in os.walk(source_path):
            for file in files:
                src_file = os.path.join(root, file)
                shutil.copy(src_file, target_path)
                #file_count += 1
                #print(src_file)
    #return int(file_count)

def Match_Address(data):
    PATTERN1 = r'([\u4e00-\u9fa5]{1,7}?(?:区|县|州)){0,1}' \
               r'([\u4e00-\u9fa5]{1,7}?(?:镇|街道办事处|乡)){0,1}' \
               r'([\u4e00-\u9fa5]{1,7}?(?:村|社区居委会|街|路|庙|坡|路))' \
               r'(((.){1,15})?(?:社|号|(.)))' \
    # \u4e00-\u9fa5 匹配任何中文
    # {2,5} 匹配2到5次
    # ? 前面可不匹配
    # (?:pattern) 如industr(?:y|ies) 就是一个比 'industry|industries' 更简略的表达式。意思就是说括号里面的内容是一个整体是以y或者ies结尾的单词
    pattern = re.compile(PATTERN1)
    p1 = ''
    p2 = ''
    p3 = ''
    p4 = ''

    m = pattern.search(data)
    if not m:
        return data
    path = ''
    #区|县|州
    if m.lastindex >= 1:
        p1 = m.group(1)
        if p1 is not None:
            path = path + p1 + "\\"
    #镇|街道办事处
    if m.lastindex >= 2:
        temp = m.group(2)
        if temp is not None:
            sheIndex = temp.find('办事处')
            if sheIndex >1:
                conpileText = re.compile('街道办事处')
                p2 = conpileText.sub('镇', temp)
            else:
                p2 = temp
        path =  path + p2 + "\\"
    #村|社区居委会|街
    if m.lastindex >= 3:
        p3 = m.group(3)
        if p3 is not None:
            path =  path + p3 + "\\"
    #社|号
    if m.lastindex >= 4:
        p4 = m.group(4)
        if p4 is not None:
            path =  path + p4 + "\\"
    #path = p1 + "\\" + p2 + "\\" + p3 + "\\" + p4
    #out = '%s|%s|%s|%s|%s' % (p1, p2, p3, p4, p5)
    return path

def get_data_from_excel(excel_dir):#读取excel，取出所有sheet要执行的接口信息，返回列表
    # 打开文件
    workbook = load_workbook(excel_dir)
    # 获取所有sheet
    all_sheets = workbook.sheetnames

    # 创建工作簿
    book = xlwt.Workbook()
    indexTemp = excel_dir.rfind('.')
    book_name = excel_dir[:indexTemp]+ '[整理]'+ excel_dir[indexTemp:len(excel_dir)-1:1]
    # book_name = '2005-2017年20201215[整理].xlsx'
    #循环所有的sheet
    for i in range(0, len(all_sheets)):
        #获取当前sheet
        work_sheet = all_sheets[i]
        sheet = workbook[work_sheet]
        rows = sheet.max_row
        cols = sheet.max_column

        # 创建表单
        book.encoding = "utf-8"  # Equals to workbook = xlwt.Workbook(encoding="utf-8")
        newSheet = book.add_sheet(sheetname = work_sheet, cell_overwrite_ok=True)
        style = xlwt.XFStyle()  # 初始化样式
        font = xlwt.Font()  # 为样式创建字体
        font.name = '宋体'
        font.bold = True
        style.font = font  # 为样式设置字体

        column0 = ["档案号", "原档案号", "登记类别", "现所有权人", "原所有权人", "土地房屋坐落","登记日期","关联号" ,"页数"]
        k=0
        for k in range(0, len(column0)):
            newSheet.write(0, k, column0[k], style)

        for r in range(1, rows):  # 从2行开始取数据
            oneSheetList = ExcelData()
            #0 档案号
            fileNumber = sheet[str(r + 1)][0].value
            if fileNumber is not None:
                newSheet.write(r, 0, fileNumber, style)
            else:
                continue

            #1 原档案号
            origFileNumber = sheet[str(r + 1)][1].value
            if origFileNumber is not None:
                newSheet.write(r, 1, origFileNumber, style)

            #2 登记类别
            registCategory = sheet[str(r + 1)][2].value
            if registCategory is not None:
                newSheet.write(r, 2, registCategory, style)

            #3 权利人
            holdString = sheet[str(r + 1)][3].value
            print(f'{work_sheet},{holdString},{r}')
            if holdString is not None:
                oneSheetList.segHolders(holdString)
                newSheet.write(r, 3, oneSheetList.newHolder, style)
                newSheet.write(r, 4, oneSheetList.origiHolderList, style)

            # 4 坐落
            landHouseLocal = sheet[str(r + 1)][4].value
            if landHouseLocal is not None:
                newSheet.write(r, 5, landHouseLocal, style)

            #5 登记日期
            registDate = sheet[str(r + 1)][5].value
            if registDate is not None:
                newSheet.write(r, 6, registDate, style)
            # registDate = sheet[str(r + 1)][5].value
            # if(type(registDate)==str):
            #     registDate1 = datetime.date(registDate)  # date()内参数需要datetime.datetime型
            #     oneSheetList.registDate.append(registDate1)
            # else:
            #   oneSheetList.registDate.append(registDate)


            #6 关联号
            associat = sheet[str(r + 1)][6].value
            if associat is not None:
                associatList = []
                associatList = associat.split('\n')
                newSheet.write(r, 7, associat, style)            #

            #7 页数
            pageNumber = sheet[str(r + 1)][7].value
            if pageNumber is not None:
                newSheet.write(r, 8, pageNumber, style)

            #根据关联号按照村落排序
            #分割关联号
            listNum = 0
            if associatList is not None:
                while listNum < len(associatList):
                   yearIndex = associatList[listNum].find("C81")+4
                   endIndex = associatList[listNum].rfind("-")
                   year = associatList[listNum][yearIndex:endIndex:1]
                   associatNum = associatList[listNum][:22:1]
                   listNum = listNum + 1
                   tempNum = 0
                   while tempNum< len(root):
                       srcpath = root[tempNum] + "\\" + '_' + year + "\\" + associatNum
                       tempNum = tempNum + 1
                       #判断是否存在文文件路径
                       if os.path.exists(srcpath):
                            #拼接输出路径
                            Path = Match_Address(landHouseLocal)
                            outPathTemp = outroot + "\\" + Path + oneSheetList.newHolder + associatNum
                            sheIndex = outPathTemp.find(' ')
                            if sheIndex > 0:
                                conpileText = re.compile(' ')
                                outPath = conpileText.sub('', outPathTemp)
                            else:
                                outPath = outPathTemp
                            print(outPath)
                            #将当前文件夹所有文件复制到指定目录下
                            copy_dirs(srcpath, outPath)


    book.save(book_name)  # 保存工作簿
    print(f"{book_name}表格写入数据成功！")

root= []
root1 = 'G:\潼南农房数据清理\原始数据\_ftpfjyj\_000\_TDQSDJM'
root.append(root1)
root2 = 'G:\潼南农房数据清理\原始数据\_ftpfjyj\_000\_TFQSDJM'
root.append(root2)
root3 = 'G:\潼南农房数据清理\原始数据\_ftpfjyj\_000\_TFQSDYDJM'
root.append(root3)
outroot = 'G:\潼南农房数据清理\整理之后'

excel_dir = '2005-2017年20201215.xlsx'
#excel_dir = "2005-2017年20201215.xlsx"
print(get_data_from_excel(excel_dir))

